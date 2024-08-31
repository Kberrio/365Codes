import openpyxl
from openpyxl.utils import get_column_letter

def sort_excel_sheet(file_path, sheet_name, sort_column, ascending=True):
    # Load the workbook and select the sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Get the data from the sheet
    data = list(sheet.iter_rows(values_only=True))

    # Separate header and data
    header = data[0]
    data = data[1:]

    # Sort the data based on the specified column
    col_index = header.index(sort_column)
    data.sort(key=lambda x: x[col_index], reverse=not ascending)

    # Clear the sheet
    sheet.delete_rows(1, sheet.max_row)

    # Write the header back
    for col, value in enumerate(header, start=1):
        sheet.cell(row=1, column=col, value=value)

    # Write the sorted data
    for row, row_data in enumerate(data, start=2):
        for col, value in enumerate(row_data, start=1):
            sheet.cell(row=row, column=col, value=value)

    # Save the workbook
    workbook.save(file_path)

# Example usage
file_path = 'your_excel_file.xlsx'
sheet_name = 'Sheet1'
sort_column = 'Name'  # Replace with the column name you want to sort by
ascending = True  # Set to False for descending order

sort_excel_sheet(file_path, sheet_name, sort_column, ascending)
print(f"Excel sheet '{sheet_name}' in '{file_path}' has been sorted based on the '{sort_column}' column.")